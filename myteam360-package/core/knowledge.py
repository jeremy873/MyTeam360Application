"""
Knowledge Base — Document storage, text extraction, chunking, and full-text search.
The company brain that agents can query for context.
"""

import os
import re
import uuid
import json
import logging
from datetime import datetime
from .database import get_db

logger = logging.getLogger("MyTeam360.knowledge")

UPLOAD_DIR = os.getenv("UPLOAD_DIR", "data/uploads")
CHUNK_SIZE = 1000  # characters per chunk
CHUNK_OVERLAP = 200


class KnowledgeBase:
    """Document management with text extraction and full-text search."""

    def __init__(self):
        os.makedirs(UPLOAD_DIR, exist_ok=True)

    # ── Folders ──

    def create_folder(self, owner_id: str, name: str, icon: str = "📁",
                      shared: bool = False) -> dict:
        folder_id = f"fld_{uuid.uuid4().hex[:10]}"
        with get_db() as db:
            db.execute("INSERT INTO kb_folders (id, owner_id, name, icon, shared) VALUES (?,?,?,?,?)",
                       (folder_id, owner_id, name, icon, int(shared)))
        return self.get_folder(folder_id)

    def get_folder(self, folder_id: str) -> dict | None:
        with get_db() as db:
            row = db.execute("SELECT * FROM kb_folders WHERE id=?", (folder_id,)).fetchone()
            if not row:
                return None
            d = dict(row)
            d["document_count"] = db.execute(
                "SELECT COUNT(*) as c FROM kb_documents WHERE folder_id=?", (folder_id,)).fetchone()["c"]
            return d

    def list_folders(self, user_id: str) -> list:
        with get_db() as db:
            rows = db.execute("""
                SELECT f.*, (SELECT COUNT(*) FROM kb_documents WHERE folder_id=f.id) as document_count
                FROM kb_folders f WHERE f.owner_id=? OR f.shared=1 ORDER BY f.name
            """, (user_id,)).fetchall()
            return [dict(r) for r in rows]

    def delete_folder(self, folder_id: str) -> bool:
        with get_db() as db:
            return db.execute("DELETE FROM kb_folders WHERE id=?", (folder_id,)).rowcount > 0

    # ── Documents ──

    def add_document(self, owner_id: str, filename: str, content: bytes,
                     folder_id: str = None, tags: list = None, shared: bool = False) -> dict:
        doc_id = f"doc_{uuid.uuid4().hex[:12]}"
        file_ext = os.path.splitext(filename)[1].lower()
        storage_path = os.path.join(UPLOAD_DIR, f"{doc_id}{file_ext}")

        # Save file
        with open(storage_path, "wb") as f:
            f.write(content)

        file_size = len(content)
        tags_json = json.dumps(tags or [])

        with get_db() as db:
            db.execute("""
                INSERT INTO kb_documents (id, folder_id, owner_id, filename, file_type, file_size,
                                          storage_path, tags, shared, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'processing')
            """, (doc_id, folder_id, owner_id, filename, file_ext, file_size,
                  storage_path, tags_json, int(shared)))

        # Extract text and create chunks
        try:
            text = self._extract_text(storage_path, file_ext)
            if text:
                chunks = self._chunk_text(text)
                page_count = text.count("\f") + 1  # rough page estimate
                with get_db() as db:
                    for i, chunk in enumerate(chunks):
                        chunk_id = f"chk_{uuid.uuid4().hex[:10]}"
                        db.execute("""
                            INSERT INTO kb_chunks (id, document_id, chunk_index, content)
                            VALUES (?, ?, ?, ?)
                        """, (chunk_id, doc_id, i, chunk))
                        # Add to FTS index
                        db.execute("""
                            INSERT INTO kb_search (chunk_id, content, document_name, tags)
                            VALUES (?, ?, ?, ?)
                        """, (chunk_id, chunk, filename, " ".join(tags or [])))

                    db.execute("""
                        UPDATE kb_documents SET status='ready', page_count=?, processed_at=CURRENT_TIMESTAMP
                        WHERE id=?
                    """, (page_count, doc_id))
                logger.info(f"Indexed document {filename}: {len(chunks)} chunks")
            else:
                with get_db() as db:
                    db.execute("UPDATE kb_documents SET status='ready', processed_at=CURRENT_TIMESTAMP WHERE id=?",
                               (doc_id,))
        except Exception as e:
            logger.error(f"Error processing {filename}: {e}")
            with get_db() as db:
                db.execute("UPDATE kb_documents SET status='error' WHERE id=?", (doc_id,))

        return self.get_document(doc_id)

    def get_document(self, doc_id: str) -> dict | None:
        with get_db() as db:
            row = db.execute("SELECT * FROM kb_documents WHERE id=?", (doc_id,)).fetchone()
            if not row:
                return None
            d = dict(row)
            try:
                d["tags"] = json.loads(d.get("tags", "[]"))
            except Exception:
                d["tags"] = []
            d["chunk_count"] = db.execute(
                "SELECT COUNT(*) as c FROM kb_chunks WHERE document_id=?", (doc_id,)).fetchone()["c"]
            return d

    def list_documents(self, user_id: str, folder_id: str = None) -> list:
        with get_db() as db:
            if folder_id:
                rows = db.execute("""
                    SELECT * FROM kb_documents WHERE (owner_id=? OR shared=1) AND folder_id=?
                    ORDER BY created_at DESC
                """, (user_id, folder_id)).fetchall()
            else:
                rows = db.execute("""
                    SELECT * FROM kb_documents WHERE owner_id=? OR shared=1
                    ORDER BY created_at DESC
                """, (user_id,)).fetchall()
            result = []
            for r in rows:
                d = dict(r)
                try:
                    d["tags"] = json.loads(d.get("tags", "[]"))
                except Exception:
                    d["tags"] = []
                result.append(d)
            return result

    def delete_document(self, doc_id: str) -> bool:
        with get_db() as db:
            doc = db.execute("SELECT storage_path FROM kb_documents WHERE id=?", (doc_id,)).fetchone()
            if doc and doc["storage_path"] and os.path.exists(doc["storage_path"]):
                os.remove(doc["storage_path"])
            # Remove from FTS
            chunk_ids = db.execute("SELECT id FROM kb_chunks WHERE document_id=?", (doc_id,)).fetchall()
            for c in chunk_ids:
                db.execute("DELETE FROM kb_search WHERE chunk_id=?", (c["id"],))
            return db.execute("DELETE FROM kb_documents WHERE id=?", (doc_id,)).rowcount > 0

    # ── Search ──

    def search(self, query: str, user_id: str = None, folder_ids: list = None,
               limit: int = 10) -> list:
        """Full-text search across knowledge base documents."""
        with get_db() as db:
            # Use FTS5 for search
            try:
                fts_query = " OR ".join(query.split())
                rows = db.execute("""
                    SELECT s.chunk_id, s.content, s.document_name, s.tags,
                           rank as relevance,
                           c.document_id, c.chunk_index, c.page_number, c.section_title,
                           d.folder_id, d.owner_id, d.shared
                    FROM kb_search s
                    JOIN kb_chunks c ON s.chunk_id = c.id
                    JOIN kb_documents d ON c.document_id = d.id
                    WHERE kb_search MATCH ?
                    AND d.status='ready'
                    ORDER BY rank
                    LIMIT ?
                """, (fts_query, limit)).fetchall()
            except Exception:
                # Fallback to LIKE search
                rows = db.execute("""
                    SELECT c.id as chunk_id, c.content, d.filename as document_name,
                           d.tags, 0 as relevance,
                           c.document_id, c.chunk_index, c.page_number, c.section_title,
                           d.folder_id, d.owner_id, d.shared
                    FROM kb_chunks c
                    JOIN kb_documents d ON c.document_id = d.id
                    WHERE c.content LIKE ? AND d.status='ready'
                    LIMIT ?
                """, (f"%{query}%", limit)).fetchall()

            # Filter by access
            results = []
            for r in rows:
                if user_id and not r["shared"] and r["owner_id"] != user_id:
                    continue
                if folder_ids and r["folder_id"] not in folder_ids:
                    continue
                results.append({
                    "chunk_id": r["chunk_id"],
                    "content": r["content"],
                    "document_name": r["document_name"],
                    "document_id": r["document_id"],
                    "page_number": r["page_number"],
                    "section_title": r["section_title"],
                })
            return results

    def get_context_for_agent(self, query: str, user_id: str,
                              folder_ids: list = None, max_chunks: int = 5) -> str:
        """Get relevant knowledge base context for an agent's prompt."""
        results = self.search(query, user_id=user_id, folder_ids=folder_ids, limit=max_chunks)
        if not results:
            return ""
        parts = ["[KNOWLEDGE BASE CONTEXT]"]
        for r in results:
            source = r["document_name"]
            if r.get("page_number"):
                source += f" (page {r['page_number']})"
            parts.append(f"\n--- Source: {source} ---\n{r['content']}")
        return "\n".join(parts)

    # ── Text Extraction ──

    def _extract_text(self, path: str, file_ext: str) -> str:
        """Extract text from various file types."""
        try:
            if file_ext in (".txt", ".md", ".csv", ".json", ".py", ".js", ".html", ".css"):
                with open(path, "r", errors="ignore") as f:
                    return f.read()

            elif file_ext == ".pdf":
                try:
                    import fitz  # PyMuPDF
                    doc = fitz.open(path)
                    text = ""
                    for page in doc:
                        text += page.get_text() + "\f"
                    doc.close()
                    return text
                except ImportError:
                    logger.warning("PyMuPDF not installed — skipping PDF extraction")
                    return ""

            elif file_ext in (".docx",):
                try:
                    from docx import Document
                    doc = Document(path)
                    return "\n".join(p.text for p in doc.paragraphs)
                except ImportError:
                    logger.warning("python-docx not installed — skipping DOCX extraction")
                    return ""

            elif file_ext in (".xlsx", ".xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    text = ""
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        text += f"\n[Sheet: {sheet}]\n"
                        for row in ws.iter_rows(values_only=True):
                            text += "\t".join(str(c) if c else "" for c in row) + "\n"
                    wb.close()
                    return text
                except ImportError:
                    logger.warning("openpyxl not installed — skipping Excel extraction")
                    return ""
        except Exception as e:
            logger.error(f"Text extraction error for {path}: {e}")
        return ""

    def _chunk_text(self, text: str) -> list:
        """Split text into overlapping chunks for search indexing."""
        if not text:
            return []
        # Clean whitespace
        text = re.sub(r"\n{3,}", "\n\n", text)
        text = re.sub(r"[ \t]+", " ", text)

        chunks = []
        start = 0
        while start < len(text):
            end = start + CHUNK_SIZE
            # Try to break at paragraph or sentence boundary
            if end < len(text):
                # Look for paragraph break
                para_break = text.rfind("\n\n", start, end)
                if para_break > start + CHUNK_SIZE // 2:
                    end = para_break
                else:
                    # Look for sentence break
                    sent_break = text.rfind(". ", start, end)
                    if sent_break > start + CHUNK_SIZE // 2:
                        end = sent_break + 1

            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            start = end - CHUNK_OVERLAP if end < len(text) else end

        return chunks

    # ── Stats ──

    def get_stats(self, user_id: str = None) -> dict:
        with get_db() as db:
            if user_id:
                docs = db.execute("SELECT COUNT(*) as c FROM kb_documents WHERE owner_id=? OR shared=1",
                                  (user_id,)).fetchone()["c"]
                chunks = db.execute("""
                    SELECT COUNT(*) as c FROM kb_chunks k
                    JOIN kb_documents d ON k.document_id=d.id
                    WHERE d.owner_id=? OR d.shared=1
                """, (user_id,)).fetchone()["c"]
            else:
                docs = db.execute("SELECT COUNT(*) as c FROM kb_documents").fetchone()["c"]
                chunks = db.execute("SELECT COUNT(*) as c FROM kb_chunks").fetchone()["c"]
            folders = db.execute("SELECT COUNT(*) as c FROM kb_folders").fetchone()["c"]
            return {"documents": docs, "chunks": chunks, "folders": folders}
